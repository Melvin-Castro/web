# Proyecto/Inventario/views.py
from django.forms import ValidationError
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.shortcuts import render, redirect, get_object_or_404
from Core.models import Categoria, Producto, Pedido, Proveedores
from Core.decorators import permisos_para
from .forms import ProductoForm, CategoriaForm, PedidoForm, ActualizarEstadoPedidoForm, ProveedoresForm

# excel
import openpyxl
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime
# bd
from django.db import transaction, IntegrityError, DatabaseError

# Vista para listar productos (para vendedor y administrador)    
@login_required
def listar_productos(request):
    productos = Producto.objects.filter(estado_registro='A').order_by('id')
    categorias = Categoria.objects.all().order_by('id')
    proveedores = Proveedores.objects.all().order_by('id')
    # Obtener parámetros GET para filtrar
    categoria_id = request.GET.get('filtro_categoria')
    precio_min = request.GET.get('filtro_precio_min')
    precio_max = request.GET.get('filtro_precio_max')
    estado_stock = request.GET.get('filtro_estado_stock')
    mostrar_todos = request.GET.get('filtro_mostrar_todos') == 'on'

    if not mostrar_todos:  # Si mostrar todos no está activo
        if categoria_id:
            productos = productos.filter(categorias=categoria_id)
        if precio_min:
            productos = productos.filter(precio_venta__gte=precio_min)
        if precio_max:
            productos = productos.filter(precio_venta__lte=precio_max)
        if estado_stock:
            productos = [p for p in productos if p.estado_stock == estado_stock]
    else:
        redirect('listar_productos')

    return render(request, 'Inventario/listar_productos.html', {
        'productos': productos,
        'categorias': categorias,
        'proveedores': proveedores,
    })

# !!!!?
@login_required
def detalle_producto(request, pk):
    producto = get_object_or_404(Producto, id=pk)
    data = {
        'nombre': producto.nombre,
        'stock': producto.stock,
        'precio_compra': producto.precio_compra,
        'precio_venta': producto.precio_venta,
    }
    return JsonResponse(data)
# Validadar errores formulario producto
def validar_datos_producto(nombre, categorias_ids, proveedor_id, stock, precio_compra, precio_venta, stock_min, stock_max):
    errores = []
    if not nombre:
        errores.append('El campo nombre es obligatorio.')
    if not categorias_ids:
        errores.append('Debe seleccionar al menos una categoría.')
    if not proveedor_id:
        errores.append('Debe de seleccionar un proveedor')
    if not stock or int(stock) < 0:
        errores.append('El campo stock debe ser un numero positivo.')
    if not precio_compra or float(precio_compra) < 0:
        errores.append('El campo precio de compra debe ser un numero positivo.')
    if not precio_venta or float(precio_venta) < 0:
        errores.append('El campo precio de venta debe ser un numero positivo.')
    if not stock_min or int(stock_min) < 0:
        errores.append('El campo stock mínimo debe ser un numero positivo.')
    if not stock_max or int(stock_max) < 0:
        errores.append('El campo stock máximo debe ser un numero positivo.')
    if stock_min and stock_max and int(stock_min) > int(stock_max):
        errores.append('El campo stock mínimo no puede ser mayor que el stock máximo.')
    return errores

# Vista para crear un nuevo producto 
@permisos_para(lambda u:u.id_permisos.inventario_pro_CUD)
def crear_producto(request):
    categorias = Categoria.objects.all().order_by('id')
    proveedores = Proveedores.objects.all().order_by('id') 
    if request.method == 'POST':
        nombre_Prod_C = request.POST.get('nombre_Prod_C')
        categorias_ids_Prod_C = request.POST.getlist('categorias_Prod_C')
        proveedor_id_Prod_C = request.POST.get('proveedor_Prod_C')
        stock_Prod_C = request.POST.get('stock_Prod_C')
        precio_compra_Prod_C = request.POST.get('precio_compra_Prod_C')
        precio_venta_Prod_C = request.POST.get('precio_venta_Prod_C')
        stock_min_Prod_C = request.POST.get('stock_min_Prod_C')
        stock_max_Prod_C = request.POST.get('stock_max_Prod_C')
        
        # Validaciones manuales
        errores = validar_datos_producto(nombre_Prod_C, categorias_ids_Prod_C, proveedor_id_Prod_C, stock_Prod_C, precio_compra_Prod_C, precio_venta_Prod_C, stock_min_Prod_C, stock_max_Prod_C)
        
        if not errores:
            try:
                with transaction.atomic(): # Agregar por transaccion
                    producto = Producto(
                        nombre = nombre_Prod_C,
                        proveedor_id=proveedor_id_Prod_C,
                        stock = stock_Prod_C,
                        precio_compra = precio_compra_Prod_C,
                        precio_venta = precio_venta_Prod_C,
                        stock_min = stock_min_Prod_C,
                        stock_max = stock_max_Prod_C
                    )
                    producto.save()
                    producto.categorias.set(categorias_ids_Prod_C)
                    messages.success(request, 'Producto creado exitosamente.')
                alertar_stock_bajo(request, producto.id)    
                return redirect('listar_productos')
            except IntegrityError as e:
                messages.error(request, f'Error de integridad: {e}')
            except DatabaseError as e:
                messages.error(request, f'Error de base de datos: {e}')
            except Exception as e:
                messages.error(request, f'Ocurrió un error inesperado: {e}')
        else:
            for error in errores:
                messages.error(request, error)
        form = ProductoForm()

    return redirect('listar_productos')

# Vista para editar un producto
@permisos_para(lambda u: u.id_permisos.inventario_pro_CUD)
def editar_producto(request, pk):
    productos = Producto.objects.all().order_by('id')
    try:
        producto = get_object_or_404(Producto, pk=pk)
        categorias = Categoria.objects.all().order_by('id')
        proveedores = Proveedores.objects.all().order_by('id')
        if request.method == 'POST':
            # Recuperar los datos enviados en el formulario
            nombre_Prod_E = request.POST.get('nombre_Prod_E')
            categorias_ids_Prod_E = request.POST.getlist('categorias_Prod_E')
            proveedor_id = int(request.POST.get('proveedor_Prod_E'))
            proveedor_id_Prod_E = get_object_or_404(Proveedores, id=proveedor_id)
            stock_Prod_E = request.POST.get('stock_Prod_E')
            precio_compra_Prod_E = request.POST.get('precio_compra_Prod_E')
            precio_venta_Prod_E = request.POST.get('precio_venta_Prod_E')
            stock_min_Prod_E = request.POST.get('stock_min_Prod_E')
            stock_max_Prod_E = request.POST.get('stock_max_Prod_E')

            # Validar los datos
            errores = validar_datos_producto(nombre_Prod_E, categorias_ids_Prod_E, proveedor_id_Prod_E, stock_Prod_E, precio_compra_Prod_E, precio_venta_Prod_E, stock_min_Prod_E, stock_max_Prod_E)

            if not errores:
                try:
                    with transaction.atomic():
                        producto.nombre = nombre_Prod_E
                        producto.proveedor = proveedor_id_Prod_E
                        producto.stock = stock_Prod_E  # Cambiado para corregir la asignación
                        producto.precio_compra = precio_compra_Prod_E
                        producto.precio_venta = precio_venta_Prod_E
                        producto.stock_min = stock_min_Prod_E
                        producto.stock_max = stock_max_Prod_E
                        print(producto.nombre)
                        print("antes de clean")
                        producto.clean()  # Llamar a las validaciones del modelo
                        print("Despues del clean")
                        producto.save()
                        producto.categorias.set(categorias_ids_Prod_E)
                        messages.success(request, 'Producto modificado exitosamente.')
                    alertar_stock_bajo(request, producto.id)
                    return redirect('listar_productos')
                except ValidationError as e:
                    messages.error(request, f'Error de validación: {e}')
                    print("Error de validadcion")
                except IntegrityError as e:
                    messages.error(request, f'Error de integridad: {e}')
                    print("Error de Integridad")
                except DatabaseError as e:
                    messages.error(request, f'Error de base de datos: {e}')
                    print("Error de base de datos")
                except Exception as e:
                    messages.error(request, f'Ocurrió un error inesperado: {e}')
                    print("Otro error", e)
            else:
                for error in errores:
                    messages.error(request, error)
                    
        form = {
            'categorias': producto.categorias.values_list('id', flat=True),
        }
            
    except Producto.DoesNotExist:
        messages.error(request, 'El producto no existe.')
        return redirect('listar_productos')
    except Exception as e:
        messages.error(request, f'Ha ocurrido un error: {str(e)}')
        return redirect('listar_productos')
    return redirect('listar_productos')


# Vista para eliminar un producto
@permisos_para(lambda u: u.id_permisos.inventario_pro_CUD)
def eliminar_producto(request, pk):    
    try:
        producto = get_object_or_404(Producto, pk=pk)
        if request.method == 'POST':
            try:
                producto.estado_registro = '*'
                producto.save()
                messages.success(request, 'Producto eliminado exitosamente.')
                return redirect('listar_productos')
            except IntegrityError as e:
                messages.error(request, f'Error de integridad: {e}')
            except DatabaseError as e:
                messages.error(request, f'Error de base de datos: {e}')
                print("Error de base de datos: ", e)
            except Exception as e:
                messages.error(request, f'Ocurrió un error inesperado: {e}')
    except Producto.DoesNotExist:
        messages.error(request, 'El producto no existe.')
        return redirect('listar_productos')
    except Exception as e:
        messages.error(request, f'Ha ocurrido un error: {str(e)}')
        return redirect('listar_productos')
    return redirect('listar_productos')
    #return render(request, 'Inventario/eliminar_producto.html', {'producto': producto})


# Exportar Excel de Productos
@login_required
def exportar_productos_excel(request):
    try:
        if request.method == 'POST':
            producto_ids = request.POST.get('producto_ids').split(',')
            productos = Producto.objects.order_by('id').filter(id__in=producto_ids)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Lista de Productos'

            encabezados = ['ID', 'Nombre', 'Categoria', 'Stock', 'Precio Compra', 'Precio Venta', 'Estado Stock']
            ws.append(encabezados)

            # Estilos
            header_fill = PatternFill(start_color="004080", end_color="004080", fill_type="solid")  # Azul oscuro
            header_font = Font(bold=True, color="FFFFFF")  # Blanco
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
            center_alignment = Alignment(horizontal="center")
            right_alignment = Alignment(horizontal="right")
            
            for cell in ws["1:1"]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_alignment
                
            # Escribir datos
            for row in productos:
                categorias = ', '.join([str(c) for c in row.categorias.all()])
                ws.append([
                    row.id,
                    row.nombre,
                    categorias,
                    row.stock,
                    row.precio_compra,
                    row.precio_venta,
                    row.estado_stock
                ])

            # Aplicar estilos a todas las celdas de datos
            for row in ws.iter_rows(min_row=2, max_col=7, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = center_alignment if cell.column <= 3 else right_alignment
            
            # Ajustar el ancho de las columnas
            column_widths = {
                'A': 10,  # ID
                'B': 30,  # Nombre
                'C': 30,  # Categoria
                'D': 10,  # Stock
                'E': 15,  # Precio Compra
                'F': 15,  # Precio Venta
                'G': 15,  # Estado Stock
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Crear y aplicar el formato de tabla
            tab = Table(displayName="TablaProductos", ref=f"A1:G{ws.max_row}")

            # Aplicar estilo de tabla
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",  # Estilo de tabla en tonos azules
                showFirstColumn=False, 
                showLastColumn=False,
                showRowStripes=True, 
                showColumnStripes=True
            )

            ws.add_table(tab)
            
            # Ajustar texto en toda la tabla
            for row in ws.iter_rows(min_row=1, max_col=7, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = cell.alignment.copy(wrap_text=True)
                    
            # Guardar el archivo en memoria
            archivo = BytesIO()
            wb.save(archivo)
            archivo.seek(0)

            now = datetime.datetime.now()
            formatted_date = now.strftime("%Y%m%d_%H%M%S")
            filename = f'productos_{formatted_date}.xlsx'

            response = HttpResponse(
                archivo,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
            return response
    except Exception as e:
        messages.error(request, f'Ha ocurrido un error al exportar: {str(e)}')
        return redirect('listar_productos')

    return HttpResponse(status=400)

@login_required
def alertar_stock_bajo(request, id):
    producto = get_object_or_404(Producto, id=id)
    if producto.estado_stock == 'bajo':
        messages.warning(request, f"Stock Bajo en: {producto.id} - {producto.nombre}")

########### CATEGORIAS ##########

# Vista para listar categorías (para vendedor y administrador)
@login_required
def listar_categorias(request):
    categorias = Categoria.objects.all().order_by('id')
    return render(request, 'Inventario/listar_categorias.html', {'categorias': categorias})


def validar_datos_categoria(nombre, descripcion):
    errores = []
    if not nombre:
        errores.append('El campo nombre es obligatorio.')
    if not descripcion:
        errores.append('Debe agregar una descripcion a la categoria.')
    if len(descripcion) > 500:
        errores.append('La Descripcion no puede ser mayor a 500 letras.')
    print(errores)
    return errores

# Vista para crear una nueva categoría
@permisos_para(lambda u: u.id_permisos.inventario_cat_CUD)
def crear_categoria(request):
    if request.method == 'POST':
        nombre_Cat_C = request.POST.get('nombre_Cat_C')
        descripcion_Cat_C = request.POST.get('descripcion_Cat_C')

        errores = validar_datos_categoria(nombre_Cat_C, descripcion_Cat_C)
        if not errores:
            try:
                with transaction.atomic():
                    categoria = Categoria(
                        nombre=nombre_Cat_C, 
                        descripcion=descripcion_Cat_C
                    )
                    categoria.save()
                    messages.success(request, 'Categoría creada exitosamente.')
                    return redirect('listar_categorias')
            except ValidationError as e:
                messages.error(request, f'Error de validación al crear la categoría: {e}')
            except IntegrityError as e:
                messages.error(request, f'Error de integridad al crear la categoría: {e}')
            except DatabaseError as e:
                messages.error(request, f'Error de base de datos al crear la categoría: {e}')
            except Exception as e:
                messages.error(request, f'Ocurrió un error inesperado al crear la categoría: {e}')
        else:
            for error in errores:
                messages.error(request, error)

    return redirect('listar_categorias')

# Vista para editar una categoría
@permisos_para(lambda u:u.id_permisos.inventario_cat_CUD)
def editar_categoria(request, pk):
    try:
        categoria = get_object_or_404(Categoria, pk=pk)
        if request.method == 'POST':
            nombre_Cat_E = request.POST.get('nombre_Cat_E')
            descripcion_Cat_E = request.POST.get('descripcion_Cat_E')
            
            errores = validar_datos_categoria(nombre_Cat_E, descripcion_Cat_E)
            if not errores:
                try:
                    with transaction.atomic():
                        categoria.nombre = nombre_Cat_E
                        categoria.descripcion = descripcion_Cat_E
                        categoria.save()
                        messages.success(request, 'Categoría modificada exitosamente.')
                        return redirect('listar_categorias')
                except ValidationError as e:
                    messages.error(request, f'Error de validación al modificar la categoría: {e}')
                except IntegrityError as e:
                    messages.error(request, f'Error de integridad al modificar la categoría: {e}')
                except DatabaseError as e:
                    messages.error(request, f'Error de base de datos al modificar la categoría: {e}')
                except Exception as e:
                    messages.error(request, f'Ocurrió un error inesperado al modificar la categoría: {e}')
            else:
                for error in errores:
                        messages.error(request, error)
        else:
            form = {
                "nombre": categoria.nombre,
                "descripcion": categoria.descripcion            
            }
    except Categoria.DoesNotExist:
        messages.error(request, 'La categoria no existe.')
        return redirect('listar_categorias')
    except Exception as e:
        messages.error(request, f'Ha ocurrido un error: {str(e)}')
        return redirect('listar_categorias')
    return redirect('listar_categorias')

# Vista para eliminar una categoría
@permisos_para(lambda u: u.id_permisos.inventario_cat_CUD)
def eliminar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)

    if request.method == 'POST':
        try:
            with transaction.atomic():
                categoria.delete()
                messages.success(request, 'Categoría eliminada exitosamente.')
                return redirect('listar_categorias')
        except Exception as e:
            messages.error(request, f'Ocurrió un error al eliminar la categoría: {e}')
    return redirect('listar_categorias')
    #return render(request, 'Inventario/eliminar_categoria.html', {'categoria': categoria})
"""
@login_required
def listar_pedidos(request):
    if request.method == 'POST':
        form = ActualizarEstadoPedidoForm(request.POST)
        if form.is_valid():
            pedido_id = request.POST.get('pedido_id')
            pedido = get_object_or_404(Pedido, id=pedido_id)
            pedido.estado = form.cleaned_data['estado']
            pedido.save()
            return redirect('listar_pedidos')  # Asegúrate de que el nombre de tu URL coincida
    else:
        form = ActualizarEstadoPedidoForm()
    
    listar_pedidos = Pedido.objects.all()
    return render(request, 'listar_pedidos.html', {'listar_pedidos': listar_pedidos, 'form': form})

#@login_required
def crear_pedidos(request):
    if request.method == 'POST':
        form = PedidoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pedido creado exitosamente.')
            return redirect('listar_pedidos')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = PedidoForm()
    
    return render(request, 'crear_pedidos.html', {'form': form})
"""

@login_required
def crear_proveedores(request):
    if request.method == 'POST':
        form = ProveedoresForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                    messages.success(request, 'Proveedor creado exitosamente.')
                    return redirect('listar_proveedores')
            except IntegrityError as e:
                messages.error(request, f'Error de integridad: {e}')
            except DatabaseError as e:
                messages.error(request, f'Error de base de datos: {e}')
                print("Error de base de datos: ", e)
            except Exception as e:
                messages.error(request, f'Ocurrió un error inesperado: {e}')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')

    return redirect('listar_proveedores')

@login_required
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedores, pk=pk)
    print(proveedor.nombre)
    if request.method == 'POST':
        form = ProveedoresForm(request.POST, instance=proveedor)
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                    messages.success(request, 'Proveedor actualizado exitosamente.')
                    return redirect('listar_proveedores')
            except IntegrityError as e:
                messages.error(request, f'Error de integridad: {e}')
            except DatabaseError as e:
                messages.error(request, f'Error de base de datos: {e}')
                print("Error de base de datos: ", e)
            except Exception as e:
                messages.error(request, f'Ocurrió un error inesperado: {e}')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')

    return redirect('listar_proveedores')

@login_required
def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedores, pk=pk)
    print(proveedor.nombre)
    if request.method == 'POST':
        try:
            with transaction.atomic():
                proveedor.estado_registro = '*'  # Suponiendo que '*' es el estado eliminado
                proveedor.save()
                print(proveedor.estado_registro)
                messages.success(request, 'Proveedor eliminado exitosamente.')
                return redirect('listar_proveedores')
        except IntegrityError as e:
            messages.error(request, f'Error de integridad: {e}')
        except DatabaseError as e:
            messages.error(request, f'Error de base de datos: {e}')
            print("Error de base de datos: ", e)
        except Exception as e:
            messages.error(request, f'Ocurrió un error inesperado: {e}')
    print("No entro para eliminar")
    print("Método de solicitud:", request.method)
    return redirect('listar_proveedores')

@login_required
def listar_proveedores(request):
    listar_proveedores = Proveedores.objects.filter(estado_registro='A').order_by('id')
    form = ProveedoresForm()
    return render(request, 'listar_proveedores.html', {'listar_proveedores': listar_proveedores, 'form': form})

#Exportar excel proveedores
@login_required
def exportar_proveedores_excel(request):
    try:
        if request.method == 'POST':
            proveedor_ids = request.POST.get('proveedoresIds').split(',')
            print(proveedor_ids)
            proveedores = Proveedores.objects.order_by('id').filter(id__in=proveedor_ids)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Lista de Proveedores'

            encabezados = ['ID', 'Nombre', 'RUC', 'Teléfono', 'Fecha de Creación', 'Estado de Registro']
            ws.append(encabezados)

            # Estilos
            header_fill = PatternFill(start_color="004080", end_color="004080", fill_type="solid")  # Azul oscuro
            header_font = Font(bold=True, color="FFFFFF")  # Blanco
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
            center_alignment = Alignment(horizontal="center")
            right_alignment = Alignment(horizontal="right")
            
            for cell in ws["1:1"]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_alignment
                
            # Escribir datos
            for proveedor in proveedores:
                ws.append([
                    proveedor.id,
                    proveedor.nombre,
                    proveedor.ruc,
                    proveedor.telefono,
                    proveedor.fecha_creacion.strftime("%Y-%m-%d"),
                    proveedor.estado_registro
                ])

            # Aplicar estilos a todas las celdas de datos
            for row in ws.iter_rows(min_row=2, max_col=6, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = center_alignment if cell.column <= 3 else right_alignment
            
            # Ajustar el ancho de las columnas
            column_widths = {
                'A': 10,  # ID
                'B': 30,  # Nombre
                'C': 15,  # RUC
                'D': 15,  # Teléfono
                'E': 20,  # Fecha de Creación
                'F': 20,  # Estado de Registro
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Crear y aplicar el formato de tabla
            tab = Table(displayName="TablaProveedores", ref=f"A1:F{ws.max_row}")

            # Aplicar estilo de tabla
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",  # Estilo de tabla en tonos azules
                showFirstColumn=False, 
                showLastColumn=False,
                showRowStripes=True, 
                showColumnStripes=True
            )

            ws.add_table(tab)
            
            # Ajustar texto en toda la tabla
            for row in ws.iter_rows(min_row=1, max_col=6, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = cell.alignment.copy(wrap_text=True)
                    
            # Guardar el archivo en memoria
            archivo = BytesIO()
            wb.save(archivo)
            archivo.seek(0)

            now = datetime.datetime.now()
            formatted_date = now.strftime("%Y%m%d_%H%M%S")
            filename = f'proveedores_{formatted_date}.xlsx'

            response = HttpResponse(
                archivo,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
            return response
    except Exception as e:
        messages.error(request, f'Ha ocurrido un error al exportar: {str(e)}')
        return redirect('listar_proveedores')

    return HttpResponse(status=400)